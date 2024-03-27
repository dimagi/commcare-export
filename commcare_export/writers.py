import csv
import datetime
import io
import zipfile
from itertools import zip_longest

import sqlalchemy
from sqlalchemy.exc import NoSuchTableError

from alembic.migration import MigrationContext
from alembic.operations import Operations
from commcare_export.data_types import UnknownDataType, get_sqlalchemy_type
from commcare_export.specs import TableSpec
from commcare_export import get_logger

logger = get_logger(__file__)
MAX_COLUMN_SIZE = 2000


def ensure_text(v, convert_none=False):
    if v is None:
        return '' if convert_none else v

    if isinstance(v, str):
        return v
    elif isinstance(v, bytes):
        return v
    elif isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    elif isinstance(v, datetime.date):
        return v.isoformat()
    else:
        return str(v)


def to_jvalue(v):
    if v is None:
        return None

    if isinstance(v, (str, int)):
        return v
    elif isinstance(v, bytes):
        return v
    else:
        return str(v)


class TableWriter(object):
    """
    Interface for export writers: Usable in a "with" statement, and
    while open one can call write_table.

    If the implementing class does not actually need any set up, no-op
    defaults have been provided.
    """
    max_column_length = None
    support_checkpoints = False

    # set to False if writer does not support writing to the same table
    # multiple times
    supports_multi_table_write = True

    required_columns = None

    def __enter__(self):
        return self

    def write_table(self, table: TableSpec) -> None:
        raise NotImplementedError()

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass


class CsvTableWriter(TableWriter):
    supports_multi_table_write = False

    def __init__(self, file, max_column_size=MAX_COLUMN_SIZE):
        self.file = file
        self.tables = []
        self.archive = None

    def __enter__(self):
        self.archive = zipfile.ZipFile(self.file, 'w', zipfile.ZIP_DEFLATED)
        return self

    def write_table(self, table):
        if self.archive is None:
            raise Exception('Attempt to write to a closed CsvWriter')

        tempfile = io.StringIO()
        writer = csv.writer(tempfile, dialect=csv.excel)
        writer.writerow(table.headings)
        for row in table.rows:
            writer.writerow(row)

        # TODO: make this a polite zip and put everything in a subfolder
        #       with the same basename as the zipfile
        self.archive.writestr(
            '%s.csv' % self.zip_safe_name(table.name),
            tempfile.getvalue().encode('utf-8')
        )

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.archive.close()

    def zip_safe_name(self, name):
        return name[:31]


class Excel2007TableWriter(TableWriter):
    max_table_name_size = 31

    def __init__(self, file):
        try:
            import openpyxl
        except ImportError:
            raise Exception(
                "It doesn't look like this machine is configured for "
                "Excel export. To export to Excel you have to run the "
                "command:  pip install openpyxl"
            )

        self.file = file
        self.book = openpyxl.workbook.Workbook(write_only=True)
        self.sheets = {}

    def __enter__(self):
        return self

    def write_table(self, table):
        sheet = self.get_sheet(table)
        for row in table.rows:
            sheet.append([ensure_text(v) for v in row])

    def get_sheet(self, table):
        name = table.name
        if name not in self.sheets:
            sheet = self.book.create_sheet()
            sheet.title = name[:self.max_table_name_size]
            sheet.append([ensure_text(v) for v in table.headings])
            self.sheets[name] = sheet

        return self.sheets[name]

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.book.save(self.file)


class Excel2003TableWriter(TableWriter):
    max_table_name_size = 31

    def __init__(self, file):
        try:
            import xlwt
        except ImportError:
            raise Exception(
                "It doesn't look like this machine is configured for "
                "excel export. To export to excel you have to run the "
                "command:  pip install xlwt"
            )

        self.file = file
        self.book = xlwt.Workbook()
        self.sheets = {}

    def __enter__(self):
        return self

    def write_table(self, table):
        sheet, current_row = self.get_sheet(table)
        for row in table.rows:
            for colnum, val in enumerate(row):
                sheet.write(current_row, colnum, ensure_text(val))
            current_row += 1

        self.sheets[table.name] = (sheet, current_row)

    def get_sheet(self, table):
        name = table.name
        if name not in self.sheets:
            sheet = self.book.add_sheet(name[:self.max_table_name_size])
            sheet.title = name[:self.max_table_name_size]

            for colnum, val in enumerate(table.headings):
                sheet.write(0, colnum, ensure_text(val))

            self.sheets[name] = (sheet, 1)  # start from row 1

        return self.sheets[name]

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.book.save(self.file)


class JValueTableWriter(TableWriter):
    """
    Write tables to JSON-friendly in-memory values
    """

    def __init__(self):
        self.tables = {}

    def write_table(self, table):
        if table.name not in self.tables:
            self.tables[table.name] = TableSpec(
                name=table.name,
                headings=list(table.headings),
                rows=[],
            )
        else:
            assert self.tables[table.name].headings == list(table.headings)

        self.tables[table.name].rows = list(
            self.tables[table.name].rows
        ) + [[to_jvalue(v) for v in row] for row in table.rows]


class StreamingMarkdownTableWriter(TableWriter):
    """
    Writes markdown to an output stream, where each table just comes one
    after the other
    """
    supports_multi_table_write = False

    def __init__(self, output_stream, compute_widths=False):
        self.output_stream = output_stream
        self.compute_widths = compute_widths

    def write_table(self, table):
        col_widths = None
        if self.compute_widths:
            col_widths = self._get_column_widths(table)
            row_template = ' | '.join([
                '{{:<{}}}'.format(width) for width in col_widths
            ])
        else:
            row_template = ' | '.join(['{}'] * len(table.headings))

        if table.name:
            self.output_stream.write('\n# %s \n\n' % table.name)

        self.output_stream.write(
            '| %s |\n' % row_template.format(*table.headings)
        )
        if col_widths:
            self.output_stream.write(
                '| %s |\n'
                % row_template.format(*['-' * width for width in col_widths])
            )

        for row in table.rows:
            text_row = (ensure_text(val, convert_none=True) for val in row)
            self.output_stream.write(
                '| %s |\n' % row_template.format(*text_row)
            )

    def _get_column_widths(self, table):
        all_rows = [table.headings] + table.rows
        columns = list(map(list, zip(*all_rows)))
        col_widths = map(len, [max(col, key=len) for col in columns])
        return list(col_widths)


class SqlMixin(object):
    """
    Write tables to a database specified by URL
    (TODO) with "upsert" based on primary key.
    """

    MIN_VARCHAR_LEN = 32
    # Arbitrary point at which we switch to TEXT; for Postgres
    # VARCHAR == TEXT anyhow
    MAX_VARCHAR_LEN = 255

    def __init__(self, db_url, poolclass=None, engine=None):
        self.db_url = db_url
        self.collation = 'utf8mb4_unicode_ci' if 'mysql' in db_url else None
        self.engine = engine or sqlalchemy.create_engine(
            db_url, poolclass=poolclass
        )
        self._metadata = None

    def __enter__(self):
        self.connection = self.engine.connect()
        return self  # TODO: fork the writer so this can be called many times

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.connection.close()

    @property
    def is_postgres(self):
        return 'postgres' in self.db_url

    @property
    def is_oracle(self):
        return 'oracle' in self.db_url

    @property
    def is_mysql(self):
        return 'mysql' in self.db_url

    @property
    def is_mssql(self):
        return 'mssql' in self.db_url

    @property
    def max_column_length(self):
        if self.is_postgres:
            # https://www.postgresql.org/docs/current/static/sql-syntax-lexical.html#SQL-SYNTAX-IDENTIFIERS
            return 63
        if self.is_mysql:
            # https://dev.mysql.com/doc/refman/8.0/en/identifiers.html
            return 64
        if self.is_mssql:
            # https://docs.microsoft.com/en-us/sql/relational-databases/databases/database-identifiers?view=sql-server-2017
            return 128
        if self.is_oracle:
            return 128
        raise Exception("Unknown database dialect: {}".format(self.db_url))

    @property
    def metadata(self):
        if (
            self._metadata is None
            or self._metadata.bind.closed
            or self._metadata.bind.invalidated
        ):
            if self.connection.closed:
                raise Exception('Tried to bind to a closed connection')
            if self.connection.invalidated:
                raise Exception('Tried to bind to an invalidated connection')
            self._metadata = sqlalchemy.MetaData(bind=self.connection)
        return self._metadata

    def get_table(self, table_name):
        try:
            return sqlalchemy.Table(
                table_name,
                self.metadata,
                autoload_with=self.connection,
            )
        except NoSuchTableError:
            return None

    def get_id_column(self):
        return sqlalchemy.Column(
            'id', sqlalchemy.Unicode(self.MAX_VARCHAR_LEN), primary_key=True
        )


class SqlTableWriter(SqlMixin, TableWriter):
    """
    Write tables to a database specified by URL
    (TODO) with "upsert" based on primary key.
    """
    support_checkpoints = True
    required_columns = ['id']

    def __init__(self, db_url, strict_types=False, poolclass=None):
        super(SqlTableWriter, self).__init__(db_url, poolclass=poolclass)
        self.strict_types = strict_types

    def get_data_type(self, explicit_type, val):
        if explicit_type:
            return self.get_explicit_type(explicit_type)
        return self.best_type_for(val)

    def get_explicit_type(self, data_type):
        try:
            return get_sqlalchemy_type(data_type)
        except UnknownDataType:
            if data_type:
                logger.warning(
                    "Found unknown data type '{data_type}'".format(
                        data_type=data_type,
                    )
                )
            return self.best_type_for('')  # todo: more explicit fallback

    def best_type_for(self, val):
        if isinstance(val, bool):
            return sqlalchemy.Boolean()
        elif isinstance(val, datetime.datetime):
            if self.is_mssql:
                return sqlalchemy.dialects.mssql.DATETIME2()
            else:
                return sqlalchemy.DateTime()
        elif isinstance(val, datetime.date):
            return sqlalchemy.Date()

        if isinstance(val, int):
            return sqlalchemy.Integer()
        elif isinstance(val, str):
            if self.is_postgres:
                # PostgreSQL is the best; you can use TEXT everywhere
                # and it works like a charm.
                return sqlalchemy.UnicodeText(collation=self.collation)
            elif self.is_mysql:
                # MySQL cannot build an index on TEXT due to the lack of
                # a field length, so we try to use VARCHAR when
                # possible.
                if len(val) < self.MAX_VARCHAR_LEN:
                    return sqlalchemy.Unicode(
                        max(len(val), self.MIN_VARCHAR_LEN),
                        collation=self.collation
                    )
                else:
                    return sqlalchemy.UnicodeText(collation=self.collation)
            elif self.is_mssql:
                # MSSQL (pre 2016) doesn't allow indices on columns
                # longer than 900 bytes
                # https://docs.microsoft.com/en-us/sql/t-sql/statements/create-index-transact-sql
                # If any of our data is bigger than this, then set the
                # column to NVARCHAR(max) `length` here is the size in
                # bytes
                # https://docs.sqlalchemy.org/en/13/core/type_basics.html#sqlalchemy.types.String.params.length
                length_in_bytes = len(val.encode('utf-8'))
                column_length_in_bytes = None if length_in_bytes > 900 else 900
                return sqlalchemy.NVARCHAR(
                    length=column_length_in_bytes, collation=self.collation
                )
            elif self.is_oracle:
                return sqlalchemy.Unicode(4000, collation=self.collation)
            else:
                raise Exception(f"Unknown database dialect: {self.db_url}")
        else:
            # We do not have a name for "bottom" in SQL aka the type
            # whose least upper bound with any other type is the other
            # type.
            return sqlalchemy.UnicodeText(collation=self.collation)

    def compatible(self, source_type, dest_type):
        """
        Checks _coercion_ compatibility.
        """
        if isinstance(source_type, sqlalchemy.String):
            if not isinstance(dest_type, sqlalchemy.String):
                return False
            elif source_type.length is None:
                # The length being None means that we are looking at
                # indefinite strings aka TEXT. This tool will never
                # create strings with bounds, but if a target DB has one
                # then we cannot insert to it. We will request that
                # whoever uses this tool convert to TEXT type.
                return dest_type.length is None
            else:
                return dest_type.length is None or (
                    dest_type.length >= source_type.length
                )

        compatibility = {
            sqlalchemy.String: (sqlalchemy.Text,),
            sqlalchemy.Integer: (sqlalchemy.String, sqlalchemy.Text),
            sqlalchemy.Boolean:
                (sqlalchemy.String, sqlalchemy.Text, sqlalchemy.Integer),
            sqlalchemy.DateTime:
                (sqlalchemy.String, sqlalchemy.Text, sqlalchemy.Date),
            sqlalchemy.Date: (sqlalchemy.String, sqlalchemy.Text),
        }

        # add dialect specific types
        try:
            compatibility[sqlalchemy.JSON
                         ] = (sqlalchemy.dialects.postgresql.json.JSON,)
        except AttributeError:
            pass
        try:
            compatibility[sqlalchemy.Boolean
                         ] += (sqlalchemy.dialects.mssql.base.BIT,)
        except AttributeError:
            pass

        for _type, types in compatibility.items():
            if isinstance(source_type, _type):
                return isinstance(dest_type, (_type,) + types)

    def strict_types_compatibility_check(self, source_type, dest_type, val):
        if isinstance(source_type, sqlalchemy.String):
            if not isinstance(dest_type, sqlalchemy.String):
                return  # Can't do anything
            elif dest_type.length is None:
                return  # already a TEXT column
            elif isinstance(val, str) and dest_type.length >= len(val):
                return  # no need to upgrade to TEXT column
            elif source_type.length is None:
                return sqlalchemy.UnicodeText(collation=self.collation)
            elif dest_type.length < source_type.length:
                return source_type

    def least_upper_bound(self, source_type, dest_type):
        """
        Returns the _coercion_ least upper bound.

        Mostly just promotes everything to string if it is not already.
        In fact, since this is only called when they are incompatible,
        it promotes to string right away.
        """

        # FIXME: Don't be so silly
        return sqlalchemy.UnicodeText(collation=self.collation)

    def make_table_compatible(self, table, row_dict, data_type_dict):
        ctx = MigrationContext.configure(self.connection)
        op = Operations(ctx)
        columns = {c.name: c for c in table.columns}

        for column, val in row_dict.items():
            if val is None:
                continue

            val_type = self.get_data_type(data_type_dict[column], val)
            if column not in columns:
                logger.warning(
                    f"Adding column '{table.name}.{column} {val_type}'"
                )
                op.add_column(
                    table.name,
                    sqlalchemy.Column(column, val_type, nullable=True)
                )
                self.metadata.clear()
                table = self.get_table(table.name)
            elif not columns[column].primary_key:
                col_type = columns[column].type
                new_col_type = None
                if self.strict_types:
                    # don't bother checking compatibility since we're
                    # not going to change anything
                    new_col_type = self.strict_types_compatibility_check(
                        val_type, col_type, val
                    )
                elif not self.compatible(val_type, col_type):
                    new_col_type = self.least_upper_bound(val_type, col_type)

                if new_col_type:
                    logger.warning(
                        f'Altering column {columns[column]} from {col_type} '
                        f'to {new_col_type} for value: "{type(val)}:{val}"',
                    )
                    op.alter_column(table.name, column, type_=new_col_type)
                    self.metadata.clear()
                    table = self.get_table(table.name)
        return table

    def create_table(self, table_name, row_dict, data_type_dict):
        ctx = MigrationContext.configure(self.connection)
        op = Operations(ctx)
        if self.strict_types:
            create_sql = sqlalchemy.schema.CreateTable(
                sqlalchemy.Table(
                    table_name,
                    sqlalchemy.MetaData(),
                    *self._get_columns_for_data(row_dict, data_type_dict)
                )
            ).compile(self.connection.engine)
            logger.warning(
                f"Table '{table_name}' does not exist. Creating table "
                f"with:\n{create_sql}"
            )
            empty_cols = [
                name for (name, val) in row_dict.items()
                if val is None and name not in data_type_dict
            ]
            if empty_cols:
                logger.warning(
                    "This schema does not include the following columns "
                    "since we are unable to determine the column type at "
                    f"this stage: {empty_cols}"
                )
        op.create_table(
            table_name,
            *self._get_columns_for_data(row_dict, data_type_dict)
        )
        self.metadata.clear()
        return self.get_table(table_name)

    def upsert(self, table, row_dict):
        # For atomicity "insert, catch, update" is slightly better than
        # "select, insert or update". The latter may crash, while the
        # former may overwrite data (which should be fine if whatever is
        # racing against this is importing from the same source... if
        # not you are busted anyhow

        # strip out values that are None since the column may not exist
        # yet
        row_dict = {
            col: val for col, val in row_dict.items() if val is not None
        }
        try:
            insert = table.insert().values(**row_dict)
            self.connection.execute(insert)
        except sqlalchemy.exc.IntegrityError:
            update = (table.update()
                      .where(table.c.id == row_dict['id'])
                      .values(**row_dict))
            self.connection.execute(update)

    def write_table(self, table_spec: TableSpec) -> None:
        table_name = table_spec.name
        headings = table_spec.headings
        data_type_dict = dict(zip_longest(headings, table_spec.data_types))
        for i, row in enumerate(table_spec.rows):
            row_dict = dict(zip(headings, row))
            if i == 0:
                table = self.get_table(table_name)
                if table is None:
                    table = self.create_table(
                        table_name,
                        row_dict,
                        data_type_dict,
                    )
            # Checks the data type for every cell in every row. Maybe we
            # can use a future version of the data dictionary to avoid
            # this?
            table = self.make_table_compatible(table, row_dict, data_type_dict)
            self.upsert(table, row_dict)

    def _get_columns_for_data(self, row_dict, data_type_dict):
        return [self.get_id_column()] + [
            sqlalchemy.Column(
                column_name,
                self.get_data_type(data_type_dict[column_name], val),
                nullable=True
            )
            for (column_name, val) in row_dict.items()
            if ((val is not None or data_type_dict[column_name])
                and column_name != 'id')
        ]

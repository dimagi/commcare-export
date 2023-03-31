import os.path
import pprint
import unittest

import openpyxl

from commcare_export.builtin_queries import ColumnEnforcer
from commcare_export.env import BuiltInEnv, JsonPathEnv
from commcare_export.excel_query import *
from commcare_export.excel_query import _get_safe_source_field


class TestExcelQuery(unittest.TestCase):

    @classmethod
    def setup_class(cls):
        pass

    def test_split_leftmost(self):
        assert split_leftmost(
            parse_jsonpath('foo')
        ) == (jsonpath.Fields('foo'), jsonpath.This())
        assert split_leftmost(
            parse_jsonpath('foo.baz')
        ) == (jsonpath.Fields('foo'), jsonpath.Fields('baz'))
        assert split_leftmost(parse_jsonpath('foo.baz.bar')) == (
            jsonpath.Fields('foo'),
            jsonpath.Fields('baz').child(jsonpath.Fields('bar'))
        )
        assert split_leftmost(
            parse_jsonpath('[*].baz')
        ) == (jsonpath.Slice(), jsonpath.Fields('baz'))
        assert split_leftmost(parse_jsonpath('foo[*].baz')) == (
            jsonpath.Fields('foo'),
            jsonpath.Slice().child(jsonpath.Fields('baz'))
        )

    def test_get_safe_source_field(self):
        assert _get_safe_source_field(
            'foo.bar.baz') == Reference('foo.bar.baz')
        assert _get_safe_source_field('foo[*].baz') == Reference('foo[*].baz')
        assert _get_safe_source_field(
            'foo..baz[*]') == Reference('foo..baz[*]')
        assert _get_safe_source_field('foo.#baz') == Reference('foo."#baz"')
        assert _get_safe_source_field(
            'foo.bar[*]..%baz') == Reference('foo.bar[*].."%baz"')
        assert _get_safe_source_field(
            'foo.bar:1.baz') == Reference('foo."bar:1".baz')

        try:
            assert _get_safe_source_field('foo.bar.')
            assert False, "Expected exception"
        except Exception:
            pass

    def test_compile_mappings(self):
        test_cases = [
            (
                'mappings.xlsx', {
                    'a': {
                        'w': 12,
                        'x': 13,
                        'y': 14,
                        'z': 15,
                        'q': 16,
                        'r': 17,
                    },
                    'b': {
                        'www': 'hello',
                        'xxx': 'goodbye',
                        'yyy': 'what is up',
                    },
                    'c': {
                        1: 'foo',
                        2: 'bar',
                        3: 'biz',
                        4: 'bizzle',
                    }
                }
            ),
        ]

        def flatten(dd):
            if isinstance(dd, defaultdict):
                return dict([(key, flatten(val)) for key, val in dd.items()])
            else:
                return dd

        for filename, mappings in test_cases:
            abs_path = os.path.join(os.path.dirname(__file__), filename)
            compiled = compile_mappings(
                openpyxl.load_workbook(abs_path)['Mappings']
            )
            # Print will be suppressed by pytest unless it fails
            if not (flatten(compiled) == mappings):
                print('In %s:' % filename)
                pprint.pprint(flatten(compiled))
                print('!=')
                pprint.pprint(mappings)
            assert flatten(compiled) == mappings

    def test_parse_sheet(self):

        test_cases = [
            (
                '001_JustDataSource.xlsx',
                SheetParts(
                    name='Forms',
                    headings=[],
                    source=Apply(
                        Reference("api_data"), Literal("form"),
                        Reference('checkpoint_manager')
                    ),
                    body=None,
                    data_source="form"
                ),
            ),
            # (
            #     '001a_JustDataSource_LibreOffice.xlsx',
            #     Emit(
            #         table='Forms',
            #         headings=[],
            #         source=Apply(Reference("api_data"), Literal("form"))
            #     )
            # ),
            (
                '002_DataSourceAndFilters.xlsx',
                SheetParts(
                    name='Forms',
                    headings=[],
                    source=Apply(
                        Reference("api_data"), Literal("form"),
                        Reference("checkpoint_manager"),
                        Literal({
                            'app_id': ['foobizzle'],
                            'type': ['intake'],
                        })
                    ),
                    body=None,
                    data_source="form"
                )
            ),
            (
                '003_DataSourceAndEmitColumns.xlsx',
                SheetParts(
                    name='Forms',
                    headings=[
                        Literal('Form Type'),
                        Literal('Fecha de Nacimiento'),
                        Literal('Sexo'),
                        Literal('Danger 0'),
                        Literal('Danger 1'),
                        Literal('Danger Fever'),
                        Literal('Danger error'),
                        Literal('Danger error'),
                        Literal('special'),
                        Literal('Danger substring 1'),
                        Literal('Danger substring 2'),
                        Literal('Danger substring error 3'),
                        Literal('Danger substring error 4'),
                        Literal('Danger substring error 5')
                    ],
                    source=Apply(
                        Reference("api_data"), Literal("form"),
                        Reference('checkpoint_manager')
                    ),
                    body=List([
                        Reference("type"),
                        Apply(
                            Reference("FormatDate"),
                            Reference("date_of_birth")
                        ),
                        Apply(Reference("sexo"), Reference("gender")),
                        Apply(
                            Reference("selected-at"), Reference("dangers"),
                            Literal(0)
                        ),
                        Apply(
                            Reference("selected-at"), Reference("dangers"),
                            Literal(1)
                        ),
                        Apply(
                            Reference("selected"), Reference("dangers"),
                            Literal('fever')
                        ),
                        Literal(
                            'Error: selected-at index must be an integer: '
                            'selected-at(abc)'
                        ),
                        Literal('Error: Unable to parse: selected(fever'),
                        Reference('path."#text"'),
                        Apply(
                            Reference("substr"), Reference("dangers"),
                            Literal(0), Literal(10)
                        ),
                        Apply(
                            Reference("substr"), Reference("dangers"),
                            Literal(4), Literal(3)
                        ),
                        Literal(
                            'Error: both substr arguments must be '
                            'non-negative integers: substr(a, b)'
                        ),
                        Literal(
                            'Error: both substr arguments must be '
                            'non-negative integers: substr(-1, 10)'
                        ),
                        Literal(
                            'Error: both substr arguments must be '
                            'non-negative integers: substr(3, -4)'
                        )
                    ]),
                    data_source="form"
                )
            ),
            (
                '005_DataSourcePath.xlsx',
                SheetParts(
                    name='Forms',
                    headings=[],
                    source=Apply(
                        Reference("api_data"), Literal("form"),
                        Reference('checkpoint_manager')
                    ),
                    body=None,
                    root_expr=Reference(
                        'form.delivery_information.child_questions.[*]'
                    ),
                    data_source="form"
                )
            ),
            (
                '006_IncludeReferencedItems.xlsx',
                SheetParts(
                    name='Forms',
                    headings=[],
                    source=Apply(
                        Reference("api_data"), Literal("form"),
                        Reference("checkpoint_manager"), Literal(None),
                        Literal(['foo', 'bar', 'bizzle'])
                    ),
                    body=None,
                    data_source="form"
                )
            ),
            (
                '010_JustDataSourceTableName.xlsx',
                SheetParts(
                    name='my_table',
                    headings=[],
                    source=Apply(
                        Reference("api_data"), Literal("form"),
                        Reference('checkpoint_manager')
                    ),
                    body=None,
                    data_source="form"
                )
            ),
        ]

        for filename, minilinq in test_cases:
            # This output will be captured by pytest and printed in case
            # of failure; helpful to isolate which test case
            print(f'Compiling sheet {filename}')
            abs_path = os.path.join(os.path.dirname(__file__), filename)
            compiled = parse_sheet(openpyxl.load_workbook(abs_path).active)
            # Print will be suppressed by pytest unless it fails
            if not (compiled == minilinq):
                print(f'In {filename}:')
                pprint.pprint(compiled)
                print('!=')
                pprint.pprint(minilinq)
            assert compiled == minilinq

    def test_parse_workbook(self):
        field_mappings = {'t1': 'Form 1', 't2': 'Form 2'}
        test_cases = [
            (
                '004_TwoDataSources.xlsx', [
                    SheetParts(
                        name='Forms',
                        headings=[],
                        source=Apply(
                            Reference("api_data"), Literal("form"),
                            Reference('checkpoint_manager')
                        ),
                        body=None,
                        data_source="form"
                    ),
                    SheetParts(
                        name='Cases',
                        headings=[],
                        source=Apply(
                            Reference("api_data"), Literal("case"),
                            Reference('checkpoint_manager')
                        ),
                        body=None,
                        data_source="case"
                    )
                ]
            ),
            (
                '007_Mappings.xlsx', [
                    SheetParts(
                        name='Forms',
                        headings=[Literal('Form Type')],
                        source=Apply(
                            Reference("api_data"), Literal("form"),
                            Reference('checkpoint_manager')
                        ),
                        body=List([
                            compile_mapped_field(
                                field_mappings, Reference("type")
                            )
                        ]),
                        data_source="form"
                    )
                ]
            ),
        ]

        for filename, minilinq in test_cases:
            # This output will be captured by pytest and printed in case
            # of failure; helpful to isolate which test case
            print(f'Compiling workbook {filename}')
            abs_path = os.path.join(os.path.dirname(__file__), filename)
            compiled = parse_workbook(openpyxl.load_workbook(abs_path))
            # Print will be suppressed by pytest unless it fails
            if not (compiled == minilinq):
                print('In %s:' % filename)
                pprint.pprint(compiled)
                print('!=')
                pprint.pprint(minilinq)
            assert compiled == minilinq

    def test_compile_mapped_field(self):
        env = BuiltInEnv() | JsonPathEnv({'foo': {'bar': 'a', 'baz': 'b'}})
        expression = compile_mapped_field({'a': 'mapped from a'},
                                          Reference('foo.bar'))
        assert expression.eval(env) == 'mapped from a'

        expression = compile_mapped_field({'a': 'mapped from a'},
                                          Reference('foo.baz'))
        assert list(expression.eval(env))[0].value == 'b'

        expression = compile_mapped_field({'a': 'mapped from a'},
                                          Reference('foo.boo'))
        assert list(expression.eval(env)) == []

    def test_get_queries_from_excel(self):
        minilinq = Bind(
            'checkpoint_manager',
            Apply(
                Reference('get_checkpoint_manager'), Literal("form"),
                Literal(["Forms"])
            ),
            Emit(
                table='Forms',
                missing_value='---',
                headings=[
                    Literal('Form Type'),
                    Literal('Fecha de Nacimiento'),
                    Literal('Sexo'),
                    Literal('Danger 0'),
                    Literal('Danger 1'),
                    Literal('Danger Fever'),
                    Literal('Danger error'),
                    Literal('Danger error'),
                    Literal('special'),
                    Literal('Danger substring 1'),
                    Literal('Danger substring 2'),
                    Literal('Danger substring error 3'),
                    Literal('Danger substring error 4'),
                    Literal('Danger substring error 5')
                ],
                source=Map(
                    source=Apply(
                        Reference("api_data"), Literal("form"),
                        Reference('checkpoint_manager')
                    ),
                    body=List([
                        Reference("type"),
                        Apply(
                            Reference("FormatDate"),
                            Reference("date_of_birth")
                        ),
                        Apply(Reference("sexo"), Reference("gender")),
                        Apply(
                            Reference("selected-at"), Reference("dangers"),
                            Literal(0)
                        ),
                        Apply(
                            Reference("selected-at"), Reference("dangers"),
                            Literal(1)
                        ),
                        Apply(
                            Reference("selected"), Reference("dangers"),
                            Literal('fever')
                        ),
                        Literal(
                            'Error: selected-at index must be an integer: '
                            'selected-at(abc)'
                        ),
                        Literal('Error: Unable to parse: selected(fever'),
                        Reference('path."#text"'),
                        Apply(
                            Reference("substr"), Reference("dangers"),
                            Literal(0), Literal(10)
                        ),
                        Apply(
                            Reference("substr"), Reference("dangers"),
                            Literal(4), Literal(3)
                        ),
                        Literal(
                            'Error: both substr arguments must be '
                            'non-negative integers: substr(a, b)'
                        ),
                        Literal(
                            'Error: both substr arguments must be '
                            'non-negative integers: substr(-1, 10)'
                        ),
                        Literal(
                            'Error: both substr arguments must be '
                            'non-negative integers: substr(3, -4)'
                        )
                    ])
                )
            )
        )

        self._compare_minilinq_to_compiled(
            minilinq, '003_DataSourceAndEmitColumns.xlsx'
        )

    def test_alternate_source_fields(self):
        minilinq = List([
            # First sheet uses a CSV column and also tests combining "Map Via"
            Bind(
                'checkpoint_manager',
                Apply(
                    Reference('get_checkpoint_manager'), Literal("form"),
                    Literal(["Forms"])
                ),
                Emit(
                    table='Forms',
                    missing_value='---',
                    headings=[
                        Literal('dob'),
                    ],
                    source=Map(
                        source=Apply(
                            Reference("api_data"), Literal("form"),
                            Reference('checkpoint_manager')
                        ),
                        body=List([
                            Apply(
                                Reference("str2date"),
                                Apply(
                                    Reference("or"), Reference("dob"),
                                    Reference("date_of_birth"),
                                    Reference("d_o_b")
                                )
                            ),
                        ])
                    )
                )
            ),

            # Second sheet uses multiple alternate source field columns (listed out of order)
            Bind(
                'checkpoint_manager',
                Apply(
                    Reference('get_checkpoint_manager'), Literal("form"),
                    Literal(["Forms1"])
                ),
                Emit(
                    table='Forms1',
                    missing_value='---',
                    headings=[
                        Literal('dob'),
                        Literal('Sex'),
                    ],
                    source=Map(
                        source=Apply(
                            Reference("api_data"), Literal("form"),
                            Reference('checkpoint_manager')
                        ),
                        body=List([
                            Reference("dob"),
                            Apply(
                                Reference("or"), Reference("gender"),
                                Reference("sex"), Reference("sex0")
                            )
                        ])
                    )
                )
            ),
        ])

        self._compare_minilinq_to_compiled(
            minilinq, '011_AlternateSourceFields.xlsx'
        )

    def test_columns_with_data_types(self):
        minilinq = Bind(
            'checkpoint_manager',
            Apply(
                Reference('get_checkpoint_manager'), Literal("form"),
                Literal(["Forms"])
            ),
            Emit(
                table='Forms',
                missing_value='---',
                headings=[
                    Literal('Name'),
                    Literal('Date of Birth'),
                    Literal('No Type Set'),
                    Literal('A Number'),
                    Literal('Bad Type'),
                ],
                source=Map(
                    source=Apply(
                        Reference("api_data"), Literal("form"),
                        Reference('checkpoint_manager')
                    ),
                    body=List([
                        Reference("name"),
                        Reference("date_of_birth"),
                        Reference("no_type_set"),
                        Reference("a_number"),
                        Reference("bad_type"),
                    ])
                ),
                data_types=[
                    Literal('text'),
                    Literal('date'),
                    Literal(None),
                    Literal('integer'),
                    Literal('bad_type'),
                ],
            ),
        )
        self._compare_minilinq_to_compiled(
            minilinq, '012_ColumnsWithTypes.xlsx'
        )

    def test_multi_emit(self):
        minilinq = List([
            Bind(
                "checkpoint_manager",
                Apply(
                    Reference('get_checkpoint_manager'), Literal("form"),
                    Literal(["Forms", "Cases"])
                ),
                Filter(
                    predicate=Apply(Reference("filter_empty"), Reference("$")),
                    source=Map(
                        source=Apply(
                            Reference("api_data"), Literal("form"),
                            Reference('checkpoint_manager')
                        ),
                        body=List([
                            Emit(
                                table="Forms",
                                headings=[Literal("id"),
                                          Literal("name")],
                                missing_value='---',
                                source=Map(
                                    source=Reference("`this`"),
                                    body=List([
                                        Reference("id"),
                                        Reference("form.name"),
                                    ]),
                                )
                            ),
                            Emit(
                                table="Cases",
                                headings=[Literal("case_id")],
                                missing_value='---',
                                source=Map(
                                    source=Reference("form..case"),
                                    body=List([
                                        Reference("@case_id"),
                                    ]),
                                )
                            )
                        ])
                    )
                )
            ),
            Bind(
                'checkpoint_manager',
                Apply(
                    Reference('get_checkpoint_manager'), Literal("case"),
                    Literal(["Other cases"])
                ),
                Emit(
                    table="Other cases",
                    headings=[Literal("id")],
                    missing_value='---',
                    source=Map(
                        source=Apply(
                            Reference("api_data"), Literal("case"),
                            Reference('checkpoint_manager')
                        ),
                        body=List([Reference("id")])
                    )
                )
            )
        ])

        self._compare_minilinq_to_compiled(
            minilinq, '008_multiple-tables.xlsx', combine_emits=True
        )

    def test_multi_emit_no_combine(self):
        minilinq = List([
            Bind(
                "checkpoint_manager",
                Apply(
                    Reference('get_checkpoint_manager'), Literal("form"),
                    Literal(["Forms"])
                ),
                Emit(
                    table="Forms",
                    headings=[Literal("id"), Literal("name")],
                    missing_value='---',
                    source=Map(
                        source=Apply(
                            Reference("api_data"), Literal("form"),
                            Reference('checkpoint_manager')
                        ),
                        body=List([
                            Reference("id"),
                            Reference("form.name"),
                        ]),
                    )
                )
            ),
            Bind(
                "checkpoint_manager",
                Apply(
                    Reference('get_checkpoint_manager'), Literal("form"),
                    Literal(["Cases"])
                ),
                Emit(
                    table="Cases",
                    headings=[Literal("case_id")],
                    missing_value='---',
                    source=Map(
                        source=FlatMap(
                            body=Reference("form..case"),
                            source=Apply(
                                Reference("api_data"), Literal("form"),
                                Reference('checkpoint_manager')
                            )
                        ),
                        body=List([
                            Reference("@case_id"),
                        ]),
                    )
                )
            ),
            Bind(
                "checkpoint_manager",
                Apply(
                    Reference('get_checkpoint_manager'), Literal("case"),
                    Literal(["Other cases"])
                ),
                Emit(
                    table="Other cases",
                    headings=[Literal("id")],
                    missing_value='---',
                    source=Map(
                        source=Apply(
                            Reference("api_data"), Literal("case"),
                            Reference('checkpoint_manager')
                        ),
                        body=List([Reference("id")])
                    )
                )
            )
        ])

        self._compare_minilinq_to_compiled(
            minilinq, '008_multiple-tables.xlsx', combine_emits=False
        )

    def test_multi_emit_with_organization(self):
        minilinq = List([
            Bind(
                "checkpoint_manager",
                Apply(
                    Reference('get_checkpoint_manager'), Literal("form"),
                    Literal(["Forms", "Cases"])
                ),
                Filter(
                    predicate=Apply(Reference("filter_empty"), Reference("$")),
                    source=Map(
                        source=Apply(
                            Reference("api_data"), Literal("form"),
                            Reference('checkpoint_manager')
                        ),
                        body=List([
                            Emit(
                                table="Forms",
                                headings=[
                                    Literal("id"),
                                    Literal("name"),
                                    Literal("commcare_userid")
                                ],
                                missing_value='---',
                                source=Map(
                                    source=Reference("`this`"),
                                    body=List([
                                        Reference("id"),
                                        Reference("form.name"),
                                        Reference("$.metadata.userID"),
                                    ]),
                                )
                            ),
                            Emit(
                                table="Cases",
                                headings=[
                                    Literal("case_id"),
                                    Literal("commcare_userid")
                                ],
                                missing_value='---',
                                source=Map(
                                    source=Reference("form..case"),
                                    body=List([
                                        Reference("@case_id"),
                                        Reference("$.metadata.userID"),
                                    ]),
                                )
                            )
                        ])
                    )
                )
            ),
            Bind(
                'checkpoint_manager',
                Apply(
                    Reference('get_checkpoint_manager'), Literal("case"),
                    Literal(["Other cases"])
                ),
                Emit(
                    table="Other cases",
                    headings=[Literal("id"),
                              Literal("commcare_userid")],
                    missing_value='---',
                    source=Map(
                        source=Apply(
                            Reference("api_data"), Literal("case"),
                            Reference('checkpoint_manager')
                        ),
                        body=List([Reference("id"),
                                   Reference("$.user_id")])
                    )
                )
            )
        ])

        column_enforcer = ColumnEnforcer()
        self._compare_minilinq_to_compiled(
            minilinq,
            '008_multiple-tables.xlsx',
            combine_emits=True,
            column_enforcer=column_enforcer
        )

    def test_value_or_root(self):
        minilinq = List([
            Bind(
                "checkpoint_manager",
                Apply(
                    Reference('get_checkpoint_manager'), Literal("form"),
                    Literal(["Forms"])
                ),
                Emit(
                    table="Forms",
                    headings=[Literal("id"), Literal("name")],
                    missing_value='---',
                    source=Map(
                        source=Apply(
                            Reference("api_data"), Literal("form"),
                            Reference('checkpoint_manager')
                        ),
                        body=List([
                            Reference("id"),
                            Reference("form.name"),
                        ]),
                    )
                )
            ),
            Bind(
                "checkpoint_manager",
                Apply(
                    Reference('get_checkpoint_manager'), Literal("form"),
                    Literal(["Cases"])
                ),
                Emit(
                    table="Cases",
                    headings=[Literal("case_id")],
                    missing_value='---',
                    source=Map(
                        source=FlatMap(
                            body=Apply(
                                Reference("_or_raw"), Reference("form..case"),
                                Bind(
                                    "__root_only", Literal(True),
                                    Reference("$")
                                )
                            ),
                            source=Apply(
                                Reference("api_data"), Literal("form"),
                                Reference('checkpoint_manager')
                            )
                        ),
                        body=List([
                            Reference("@case_id"),
                        ]),
                    )
                )
            ),
            Bind(
                "checkpoint_manager",
                Apply(
                    Reference('get_checkpoint_manager'), Literal("case"),
                    Literal(["Other cases"])
                ),
                Emit(
                    table="Other cases",
                    headings=[Literal("id")],
                    missing_value='---',
                    source=Map(
                        source=Apply(
                            Reference("api_data"), Literal("case"),
                            Reference('checkpoint_manager')
                        ),
                        body=List([Reference("id")])
                    )
                )
            )
        ])

        self._compare_minilinq_to_compiled(
            minilinq,
            '008_multiple-tables.xlsx',
            combine_emits=False,
            value_or_root=True
        )

    def _compare_minilinq_to_compiled(self, minilinq, filename, **kwargs):
        print("Parsing {}".format(filename))
        abs_path = os.path.join(os.path.dirname(__file__), filename)
        compiled = get_queries_from_excel(
            openpyxl.load_workbook(abs_path), missing_value='---', **kwargs
        )
        assert compiled.to_jvalue() == minilinq.to_jvalue(), filename
